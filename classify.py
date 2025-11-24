#!/usr/bin/env python3
"""
Classify bills in an Excel workbook using OpenAI's GPT-5.1 model.

For each record the script determines whether the bill is related to nutrition.
If so, it further classifies the bill's directionality, environment, and topic.

Usage:
    python classify.py path/to/bills.xlsx --sheet Sheet1
    python classify.py path/to/bills.xlsx --sheet 0 --output classified.xlsx

The script logs progress so long-running batches do not appear stalled.
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import sys
import tempfile
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook


# Classification enumerations.
DIRECTIONALITY_OPTIONS = {"neutral", "undermining", "promoting"}
ENVIRONMENT_OPTIONS = {"S_E", "FB_E", "MS_E", "PA_E", "HC_WE"}
TOPIC_OPTIONS = {
    "SchoolNutrition",
    "FoodAccess",
    "FoodAssistance",
    "Food and Beverage Taxes and Incentives",
    "FarmsAndGardens",
    "Food Labeling and Marketing",
    "Built Environment and PhysicalActivity",
    "ObesityGeneral",
    "Healthcare and Workplace Initiatives",
    "Other",
    "Institutional Standards",
}

DIRECTIONALITY_VALUES = sorted(DIRECTIONALITY_OPTIONS)
ENVIRONMENT_VALUES = sorted(ENVIRONMENT_OPTIONS)
TOPIC_VALUES = sorted(TOPIC_OPTIONS)

# Default context columns to feed into the LLM prompt (if present).
DEFAULT_CONTEXT_COLUMNS = [
    "id",
    "state",
    "summary",
    "subjects",
]

DEFAULT_MODEL = "gpt-5.1"
DEFAULT_TEMPERATURE = 0.0
DEFAULT_BATCH_SIZE: Optional[int] = None
DEFAULT_MAX_BATCH_KB = 20
DEFAULT_COMPLETION_WINDOW = "24h"
DEFAULT_POLL_INTERVAL = 15
MAX_SUMMARY_CHARS = 2500


TEXT_FORMAT = {
    "type": "json_schema",
    "name": "bill_classification_batch",
    "schema": {
        "type": "object",
        "properties": {
            "records": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "local_id": {"type": "integer"},
                        "related_to_nutrition": {
                            "type": "integer",
                            "enum": [0, 1],
                        },
                        "directionality": {
                            "type": ["string", "null"],
                            "enum": DIRECTIONALITY_VALUES + [None],
                        },
                        "environment": {
                            "type": ["string", "null"],
                            "enum": ENVIRONMENT_VALUES + [None],
                        },
                        "topic": {
                            "type": ["string", "null"],
                            "enum": TOPIC_VALUES + [None],
                        },
                    },
                    "required": [
                        "local_id",
                        "related_to_nutrition",
                        "directionality",
                        "environment",
                        "topic",
                    ],
                    "additionalProperties": False,
                },
            }
        },
        "required": ["records"],
        "additionalProperties": False,
    },
    "strict": True,
}

SYSTEM_PROMPT = """You are a policy analyst who classifies legislative bills that relate to nutrition, food systems, and physical activity.

For every record you receive:
1. Decide if the bill is related to nutrition (directly focused on nutrition, food systems, food access, food and beverage standards, marketing, or physical activity initiatives). Return `related_to_nutrition` as 1 for related bills, otherwise 0.
2. If `related_to_nutrition` is 0, output `null` for directionality, environment, and topic.
3. If related, classify the bill:
   - directionality: choose one of neutral, undermining, promoting (from the perspective of public health nutrition outcomes).
   - environment: choose one of S_E (School), FB_E (Food & Beverage), MS_E (Messaging), PA_E (Physical Activity / Built Environment), HC_WE (Health Care & Worksite).
   - topic: choose the best matching topic from SchoolNutrition, FoodAccess, FoodAssistance, Food and Beverage Taxes and Incentives, FarmsAndGardens, Food Labeling and Marketing, Built Environment and PhysicalActivity, ObesityGeneral, Healthcare and Workplace Initiatives, Other, Institutional Standards.

Prefer evidence from the summary and subjects. Use Other only when no listed topic applies. Keep answers concise and strictly follow the requested JSON schema.
"""


@dataclass
class RecordPayload:
    local_id: int
    position: int
    context: Dict[str, str]


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Classify nutrition-related attributes for bills in an Excel sheet via OpenAI."
    )
    parser.add_argument(
        "excel_path",
        help="Path to the Excel workbook containing bill records.",
    )
    parser.add_argument(
        "--sheet",
        help="Excel sheet index or name (defaults to the first sheet).",
    )
    parser.add_argument(
        "--output",
        help="Optional path for the annotated workbook. Defaults to in-place update.",
    )
    parser.add_argument(
        "--model",
        default=DEFAULT_MODEL,
        help=f"OpenAI model to use (default: {DEFAULT_MODEL}).",
    )
    parser.add_argument(
        "--api-key",
        dest="api_key",
        help="OpenAI API key. Defaults to the OPENAI_API_KEY environment variable.",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=DEFAULT_BATCH_SIZE,
        help="Maximum number of records per API request (defaults to dynamic sizing).",
    )
    parser.add_argument(
        "--temperature",
        type=float,
        default=DEFAULT_TEMPERATURE,
        help="OpenAI temperature parameter (default: 0).",
    )
    parser.add_argument(
        "--max-output-tokens",
        type=int,
        default=5000,
        help="Maximum tokens to allow in the model response (default: 5000).",
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="Optional limit on the number of records to process (useful for spot checks).",
    )
    parser.add_argument(
        "--context-columns",
        nargs="+",
        help="Override the default context columns sent to the model.",
    )
    parser.add_argument(
        "--max-batch-kb",
        type=int,
        default=DEFAULT_MAX_BATCH_KB,
        help=(
            "Approximate maximum size (in KB) of each request when --batch-size is not provided "
            f"(default: {DEFAULT_MAX_BATCH_KB})."
        ),
    )
    parser.add_argument(
        "--completion-window",
        default=DEFAULT_COMPLETION_WINDOW,
        help="Batch completion window requested from OpenAI (default: 24h).",
    )
    parser.add_argument(
        "--poll-interval",
        type=int,
        default=DEFAULT_POLL_INTERVAL,
        help="Seconds to wait between batch status polls (default: 15).",
    )
    parser.add_argument(
        "--retain-request-file",
        action="store_true",
        help="Keep the generated batch request JSONL file for debugging.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable debug logging.",
    )
    return parser


def configure_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )


def create_openai_client(api_key: Optional[str]) -> Any:
    try:
        from openai import OpenAI  
    except ImportError as exc:  # pragma: no cover - dependency missing at runtime
        raise ImportError(
            "Missing dependency 'openai'. Install it with `pip install openai`."
        ) from exc
    return OpenAI(api_key=api_key)


def parse_sheet_identifier(raw: Optional[str]) -> Optional[int | str]:
    if raw is None:
        return 0
    try:
        return int(raw)
    except (TypeError, ValueError):
        stripped = str(raw).strip()
        return stripped if stripped else 0


def resolve_sheet_name(workbook, sheet_identifier: Optional[int | str]) -> str:
    sheet_names = workbook.sheetnames
    if sheet_identifier is None:
        return sheet_names[0]
    if isinstance(sheet_identifier, int):
        return sheet_names[sheet_identifier]
    if sheet_identifier in sheet_names:
        return sheet_identifier
    lowered = sheet_identifier.lower()
    for name in sheet_names:
        if name.lower() == lowered:
            return name
    available = ", ".join(sheet_names)
    raise KeyError(f"Sheet '{sheet_identifier}' not found. Available sheets: {available}")


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def normalize_value(column: str, value: Any) -> Optional[str]:
    if is_missing(value):
        return None
    if isinstance(value, (list, tuple, set)):
        value = "; ".join(str(item) for item in value if item is not None)
    string_value = str(value).strip()
    if not string_value:
        return None
    if column == "summary" and len(string_value) > MAX_SUMMARY_CHARS:
        return f"{string_value[:MAX_SUMMARY_CHARS]}..."
    return string_value


def extract_context(
    row: pd.Series, columns: Sequence[str]
) -> Dict[str, str]:
    context: Dict[str, str] = {}
    for column in columns:
        if column not in row.index:
            continue
        value = normalize_value(column, row[column])
        if value is not None:
            context[column] = value
    return context


def build_record_payloads(
    df: pd.DataFrame, context_columns: Sequence[str], limit: Optional[int]
) -> List[RecordPayload]:
    records: List[RecordPayload] = []
    for position, (_, row) in enumerate(df.iterrows()):
        if limit is not None and position >= limit:
            break
        context = extract_context(row, context_columns)
        if not context:
            logging.debug("Record %s has no non-empty context; it will still be classified.", position)
        records.append(RecordPayload(local_id=position, position=position, context=context))
    return records


def build_user_prompt(batch: Sequence[RecordPayload]) -> str:
    lines = [
        "Classify each record below and reply with JSON that matches the provided schema.",
        "Records:",
    ]
    for record in batch:
        lines.append(f"Record {record.local_id}:")
        if record.context:
            for key, value in record.context.items():
                lines.append(f"  {key}: {value}")
        else:
            lines.append("  (No additional context provided.)")
        lines.append("")  # blank line for readability
    return "\n".join(lines).strip()


def create_batch_request_file(
    records: Sequence[RecordPayload],
    *,
    batch_size: Optional[int],
    max_batch_bytes: int,
    model: str,
    temperature: float,
    max_output_tokens: int,
) -> Tuple[Path, Dict[str, List[int]]]:
    if batch_size is not None and batch_size < 1:
        raise ValueError("Batch size must be at least 1.")
    if max_batch_bytes < 1:
        raise ValueError("max_batch_bytes must be positive.")

    def build_body(batch: Sequence[RecordPayload]) -> Dict[str, Any]:
        body: Dict[str, Any] = {
            "model": model,
            "input": [
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": build_user_prompt(batch)},
            ],
            "temperature": temperature,
            "max_output_tokens": max_output_tokens,
        }
        body["text"] = {"format": TEXT_FORMAT}
        return body

    def estimate_size(batch: Sequence[RecordPayload]) -> int:
        payload = {
            "custom_id": "size-check",
            "method": "POST",
            "url": "/v1/responses",
            "body": build_body(batch),
        }
        return len(json.dumps(payload).encode("utf-8")) + 1  # include newline

    def flush_batch(
        handle,
        batch_index: int,
        batch: Sequence[RecordPayload],
        mapping: Dict[str, List[int]],
    ) -> None:
        custom_id = f"request-{batch_index:06d}"
        mapping[custom_id] = [record.local_id for record in batch]
        request_payload = {
            "custom_id": custom_id,
            "method": "POST",
            "url": "/v1/responses",
            "body": build_body(batch),
        }
        handle.write(json.dumps(request_payload))
        handle.write("\n")

    mapping: Dict[str, List[int]] = {}
    with tempfile.NamedTemporaryFile("w", delete=False, encoding="utf-8") as handle:
        batch_index = 0
        current: List[RecordPayload] = []

        for record in records:
            current.append(record)

            if batch_size is not None and len(current) >= batch_size:
                batch_index += 1
                flush_batch(handle, batch_index, current, mapping)
                current = []
                continue

            if batch_size is None:
                current_size = estimate_size(current)
                if current_size > max_batch_bytes:
                    if len(current) == 1:
                        logging.warning(
                            "Record %s exceeds max batch size (%s bytes); sending alone.",
                            current[0].local_id,
                            max_batch_bytes,
                        )
                        batch_index += 1
                        flush_batch(handle, batch_index, current, mapping)
                        current = []
                    else:
                        overflow_record = current.pop()
                        batch_index += 1
                        flush_batch(handle, batch_index, current, mapping)
                        current = [overflow_record]

        if current:
            batch_index += 1
            flush_batch(handle, batch_index, current, mapping)

    request_path = Path(handle.name)
    logging.info(
        "Prepared %s batched requests in %s (max batch bytes: %s, enforced size: %s).",
        len(mapping),
        request_path,
        max_batch_bytes,
        "fixed" if batch_size is not None else "dynamic",
    )
    return request_path, mapping


def upload_batch_request_file(client: Any, request_file: Path) -> Any:
    with request_file.open("rb") as fh:
        uploaded = client.files.create(file=fh, purpose="batch")
    logging.info("Uploaded batch request file. file_id=%s", getattr(uploaded, "id", uploaded))
    return uploaded


def create_responses_batch(
    client: Any, *, input_file_id: str, completion_window: str
) -> Any:
    batch = client.batches.create(
        input_file_id=input_file_id,
        endpoint="/v1/responses",
        completion_window=completion_window,
    )
    logging.info(
        "Created batch %s (status=%s).",
        getattr(batch, "id", None),
        getattr(batch, "status", None),
    )
    return batch


def read_request_counts(data: Any) -> Dict[str, int]:
    raw_counts = getattr(data, "request_counts", None)
    if isinstance(raw_counts, dict):
        return raw_counts
    if hasattr(raw_counts, "model_dump"):
        return raw_counts.model_dump()
    if raw_counts is None and isinstance(data, dict):
        return data.get("request_counts") or {}
    return {}


def poll_batch_until_terminal(
    client: Any, batch_id: str, poll_interval: int
) -> Any:
    while True:
        batch = client.batches.retrieve(batch_id)
        status = getattr(batch, "status", None)
        counts = read_request_counts(batch)
        total = counts.get("total") or counts.get("submitted")
        completed = counts.get("completed") or counts.get("succeeded")
        failed = counts.get("failed") or counts.get("errored")
        logging.info(
            "Batch %s status=%s (completed=%s/%s, failed=%s).",
            batch_id,
            status,
            completed,
            total,
            failed,
        )
        if status in {"completed", "failed", "cancelled", "canceled", "expired"}:
            return batch
        time.sleep(max(1, poll_interval))


def _iter_file_bytes(stream: Any) -> bytes:
    if hasattr(stream, "iter_bytes"):
        return b"".join(stream.iter_bytes())
    if hasattr(stream, "read"):
        return stream.read()
    raise TypeError("Unexpected stream type returned from files.content.")


def download_jsonl_lines(client: Any, file_id: str) -> List[Dict[str, Any]]:
    logging.info("Downloading batch output file %s.", file_id)
    stream = client.files.content(file_id)
    payload = _iter_file_bytes(stream).decode("utf-8")
    lines: List[Dict[str, Any]] = []
    for line in payload.splitlines():
        text = line.strip()
        if not text:
            continue
        try:
            lines.append(json.loads(text))
        except json.JSONDecodeError as exc:
            logging.error("Failed to decode JSONL line from file %s: %s", file_id, exc)
    return lines


def log_batch_error_details(
    client: Any, *, batch_id: str, error_file_ids: Sequence[str]
) -> None:
    for file_id in error_file_ids:
        try:
            error_entries = download_jsonl_lines(client, file_id)
        except Exception as exc:  # noqa: BLE001
            logging.error(
                "Failed to download error file %s for batch %s: %s",
                file_id,
                batch_id,
                exc,
            )
            continue

        if not error_entries:
            logging.error(
                "Batch %s error file %s contained no details.",
                batch_id,
                file_id,
            )
            continue

        for entry in error_entries:
            logging.error(
                "Batch %s error detail: %s",
                batch_id,
                json.dumps(entry, ensure_ascii=False),
            )


def extract_response_text(response_payload: Dict[str, Any]) -> str:
    candidate: Dict[str, Any] = response_payload
    if "body" in candidate and isinstance(candidate["body"], dict):
        candidate = candidate["body"]

    output = candidate.get("output")
    if isinstance(output, list):
        for entry in output:
            content = entry.get("content")
            if isinstance(content, list):
                for part in content:
                    for key in ("text", "value"):
                        text_value = part.get(key)
                        if isinstance(text_value, str):
                            return text_value
            for key in ("output_text", "text"):
                text_value = entry.get(key)
                if isinstance(text_value, str):
                    return text_value

    if "output_text" in candidate and isinstance(candidate["output_text"], str):
        return candidate["output_text"]

    choices = candidate.get("choices")
    if isinstance(choices, list) and choices:
        message = choices[0].get("message", {})
        content = message.get("content")
        if isinstance(content, str):
            return content
        if isinstance(content, list):
            for part in content:
                if isinstance(part, dict):
                    text_value = part.get("text")
                    if isinstance(text_value, str):
                        return text_value
        if isinstance(message.get("content"), dict):
            text_value = message["content"].get("text")
            if isinstance(text_value, str):
                return text_value

    raise ValueError("Unable to locate text output in response payload.")


def parse_batch_outputs(
    outputs: List[Dict[str, Any]],
    *,
    chunk_mapping: Dict[str, List[int]],
) -> Dict[int, Tuple[int, Optional[str], Optional[str], Optional[str]]]:
    results: Dict[int, Tuple[int, Optional[str], Optional[str], Optional[str]]] = {}
    for item in outputs:
        custom_id = item.get("custom_id")
        if not custom_id:
            logging.warning("Skipping batch item without custom_id.")
            continue

        error = item.get("error")
        if error:
            logging.error("Batch item %s failed: %s", custom_id, error)
            continue

        response_payload = item.get("response")
        if not isinstance(response_payload, dict):
            logging.error("Batch item %s missing response payload.", custom_id)
            continue

        try:
            text = extract_response_text(response_payload)
        except Exception as exc:  # noqa: BLE001
            logging.error("Failed extracting response text for %s: %s", custom_id, exc)
            continue

        try:
            parsed = json.loads(text)
        except json.JSONDecodeError as exc:
            logging.error("Invalid JSON body for %s: %s\n%s", custom_id, exc, text)
            continue

        records = parsed.get("records")
        if not isinstance(records, list):
            logging.error("Response %s missing 'records' array.", custom_id)
            continue

        received_ids = set()
        for raw in records:
            try:
                normalized = normalize_classification_item(raw)
            except Exception as exc:  # noqa: BLE001
                logging.error("Invalid classification entry in %s: %s", custom_id, exc)
                continue
            local_id, related, directionality, environment, topic = normalized
            results[local_id] = (related, directionality, environment, topic)
            received_ids.add(local_id)

        expected_ids = set(chunk_mapping.get(custom_id, []))
        missing_ids = expected_ids - received_ids
        if missing_ids:
            logging.warning(
                "Batch item %s did not return results for local ids %s.",
                custom_id,
                sorted(missing_ids),
            )
    return results


def run_batch_classification(
    client: Any,
    records: Sequence[RecordPayload],
    *,
    batch_size: Optional[int],
    max_batch_bytes: int,
    model: str,
    temperature: float,
    max_output_tokens: int,
    completion_window: str,
    poll_interval: int,
    retain_request_file: bool,
) -> Dict[int, Tuple[int, Optional[str], Optional[str], Optional[str]]]:
    if not records:
        return {}

    request_file, chunk_mapping = create_batch_request_file(
        records,
        batch_size=batch_size,
        max_batch_bytes=max_batch_bytes,
        model=model,
        temperature=temperature,
        max_output_tokens=max_output_tokens,
    )

    uploaded_file = None
    batch = None
    try:
        uploaded_file = upload_batch_request_file(client, request_file)
        input_file_id = getattr(uploaded_file, "id", None)
        if not input_file_id:
            raise ValueError("OpenAI did not return an input file id for the batch.")

        batch = create_responses_batch(
            client,
            input_file_id=input_file_id,
            completion_window=completion_window,
        )
        batch_id = getattr(batch, "id", None)
        if not batch_id:
            raise ValueError("OpenAI did not return a batch id.")

        logging.info("Polling batch %s until completion...", batch_id)
        batch = poll_batch_until_terminal(
            client,
            batch_id=batch_id,
            poll_interval=poll_interval,
        )

        status = getattr(batch, "status", None)
        if status != "completed":
            raise RuntimeError(f"Batch {batch_id} ended with status '{status}'.")

        output_ids = []
        primary_output = getattr(batch, "output_file_id", None)
        if isinstance(primary_output, str):
            output_ids.append(primary_output)
        extra_outputs = getattr(batch, "output_file_ids", None)
        if isinstance(extra_outputs, list):
            output_ids.extend([value for value in extra_outputs if isinstance(value, str)])

        error_ids = []
        error_primary = getattr(batch, "error_file_id", None)
        if isinstance(error_primary, str):
            error_ids.append(error_primary)
        error_extra = getattr(batch, "error_file_ids", None)
        if isinstance(error_extra, list):
            error_ids.extend([value for value in error_extra if isinstance(value, str)])

        if not output_ids:
            if error_ids:
                log_batch_error_details(client, batch_id=batch_id, error_file_ids=error_ids)
            else:
                logging.error("Batch %s completed with no outputs and no error files.", batch_id)
            raise RuntimeError(f"Batch {batch_id} completed but no output files were provided.")

        all_outputs: List[Dict[str, Any]] = []
        for file_id in output_ids:
            all_outputs.extend(download_jsonl_lines(client, file_id))

        if error_ids:
            log_batch_error_details(client, batch_id=batch_id, error_file_ids=error_ids)

        return parse_batch_outputs(all_outputs, chunk_mapping=chunk_mapping)
    finally:
        if uploaded_file is not None:
            file_id = getattr(uploaded_file, "id", None)
            if file_id:
                try:
                    client.files.delete(file_id)
                except Exception as exc:  # noqa: BLE001
                    logging.debug("Failed to delete uploaded request file %s: %s", file_id, exc)
        if not retain_request_file:
            try:
                request_file.unlink(missing_ok=True)
            except Exception as exc:  # noqa: BLE001
                logging.debug("Failed to delete temporary request file %s: %s", request_file, exc)


def normalize_classification_item(raw: Dict[str, Any]) -> Tuple[int, int, Optional[str], Optional[str], Optional[str]]:
    if "local_id" not in raw:
        raise ValueError("Classification item missing 'local_id'.")
    local_id = int(raw["local_id"])
    related_raw = raw.get("related_to_nutrition")
    if related_raw not in (0, 1):
        raise ValueError(f"Invalid related_to_nutrition value for record {local_id}: {related_raw}")
    related = int(related_raw)

    directionality = raw.get("directionality")
    environment = raw.get("environment")
    topic = raw.get("topic")

    def cleaned_option(value: Any, valid: Sequence[str]) -> Optional[str]:
        if value is None:
            return None
        value_str = str(value).strip()
        if not value_str:
            return None
        if value_str not in valid:
            logging.warning("Unexpected value '%s'; accepted options are %s.", value_str, sorted(valid))
            return None
        return value_str

    if related == 0:
        return local_id, related, None, None, None

    directionality_clean = cleaned_option(directionality, DIRECTIONALITY_OPTIONS)
    environment_clean = cleaned_option(environment, ENVIRONMENT_OPTIONS)
    topic_clean = cleaned_option(topic, TOPIC_OPTIONS)

    return local_id, related, directionality_clean, environment_clean, topic_clean


def update_dataframe_with_results(
    df: pd.DataFrame, results: Dict[int, Tuple[int, Optional[str], Optional[str], Optional[str]]]
) -> pd.DataFrame:
    related_values: List[Optional[int]] = [None] * len(df)
    directionality_values: List[Optional[str]] = [None] * len(df)
    environment_values: List[Optional[str]] = [None] * len(df)
    topic_values: List[Optional[str]] = [None] * len(df)

    for local_id, (related, directionality, environment, topic) in results.items():
        if 0 <= local_id < len(df):
            related_values[local_id] = related
            directionality_values[local_id] = directionality
            environment_values[local_id] = environment
            topic_values[local_id] = topic
        else:
            logging.warning("Received classification for unknown record id %s; ignoring.", local_id)

    df = df.copy()
    df["related_to_nutrition"] = related_values
    df["directionality"] = directionality_values
    df["environment"] = environment_values
    df["topic"] = topic_values
    return df


def write_columns_to_workbook(
    source_path: Path,
    destination_path: Path,
    sheet_identifier: Optional[int | str],
    columns: Dict[str, Sequence[Any]],
) -> None:
    workbook = load_workbook(source_path)
    sheet_name = resolve_sheet_name(workbook, sheet_identifier)
    worksheet = workbook[sheet_name]

    header_row = next(
        worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
        None,
    )
    if not header_row:
        raise ValueError(f"Sheet '{sheet_name}' has no header row.")

    header_lookup = {
        str(value).strip().lower(): index
        for index, value in enumerate(header_row, start=1)
        if value is not None
    }

    for column_name, values in columns.items():
        key = column_name.lower()
        if key in header_lookup:
            column_index = header_lookup[key]
            logging.debug("Updating existing column '%s' (index %s).", column_name, column_index)
        else:
            column_index = (worksheet.max_column or 0) + 1
            worksheet.cell(row=1, column=column_index, value=column_name)
            header_lookup[key] = column_index
            logging.debug("Inserted new column '%s' at index %s.", column_name, column_index)

        for row_offset, value in enumerate(values, start=2):
            worksheet.cell(row=row_offset, column=column_index, value=value)

        max_row = worksheet.max_row or 1
        if len(values) + 1 < max_row:
            for row_idx in range(len(values) + 2, max_row + 1):
                worksheet.cell(row=row_idx, column=column_index, value=None)

    destination_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination_path)


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()

    configure_logging(args.verbose)
    logging.info("Starting classification run.")

    excel_path = Path(args.excel_path).expanduser().resolve()
    if not excel_path.exists():
        parser.error(f"Excel file not found: {excel_path}")

    output_path = Path(args.output).expanduser().resolve() if args.output else excel_path
    sheet_identifier = parse_sheet_identifier(args.sheet)

    try:
        logging.info("Loading sheet '%s' from %s.", args.sheet or "0", excel_path)
        df = pd.read_excel(excel_path, sheet_name=sheet_identifier)
    except Exception as exc:  # noqa: BLE001
        logging.error("Failed to read Excel sheet: %s", exc)
        raise

    if not isinstance(df, pd.DataFrame):
        parser.error("Multiple sheets were returned; please specify --sheet explicitly.")

    total_records = len(df)
    if total_records == 0:
        logging.info("No records found; nothing to classify.")
        return

    context_columns = args.context_columns or DEFAULT_CONTEXT_COLUMNS
    logging.info("Preparing %s records for classification.", total_records)
    records = build_record_payloads(df, context_columns, args.limit)
    logging.info("Prepared %s records (limit=%s).", len(records), args.limit)

    client = create_openai_client(args.api_key)

    start_time = time.perf_counter()
    results = run_batch_classification(
        client,
        records,
        batch_size=args.batch_size,
        max_batch_bytes=max(1, args.max_batch_kb) * 1024,
        model=args.model,
        temperature=args.temperature,
        max_output_tokens=args.max_output_tokens,
        completion_window=args.completion_window,
        poll_interval=max(1, args.poll_interval),
        retain_request_file=args.retain_request_file,
    )
    elapsed = time.perf_counter() - start_time
    logging.info("Completed classification for %s records in %.2f seconds.", len(results), elapsed)

    annotated_df = update_dataframe_with_results(df, results)

    columns_to_write = {
        "related_to_nutrition": annotated_df["related_to_nutrition"].tolist(),
        "directionality": annotated_df["directionality"].tolist(),
        "environment": annotated_df["environment"].tolist(),
        "topic": annotated_df["topic"].tolist(),
    }

    logging.info(
        "Writing classifications to %s%s.",
        output_path,
        " (overwriting input)" if output_path == excel_path else "",
    )
    write_columns_to_workbook(
        source_path=excel_path,
        destination_path=output_path,
        sheet_identifier=sheet_identifier,
        columns=columns_to_write,
    )
    logging.info("Classification data written successfully.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        logging.error("Classification interrupted by user.")
        sys.exit(130)
    except Exception as exc:  # noqa: BLE001
        logging.error("Classification failed: %s", exc)
        sys.exit(1)

