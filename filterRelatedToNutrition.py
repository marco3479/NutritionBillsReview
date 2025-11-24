#!/usr/bin/env python3
"""
Annotate bill records with a binary `related` flag based on nutrition keywords.

Usage:
    python filterRelatedToNutrition.py path/to/bills.xlsx
    python filterRelatedToNutrition.py path/to/bills.xlsx --output annotated.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
import re
from typing import Iterable, Sequence

import pandas as pd
from openpyxl import load_workbook

KEYWORDS = sorted(
    {
        # Query 1 - school nutrition, food access, nutrition assistance
        "school physical education",
        "school breakfasts",
        "school breakfast",
        "school lunches",
        "school lunch",
        "school meals",
        "school meal",
        "school nutrition",
        "school foods",
        "school food",
        "processed foods",
        "processed food",
        "nutrition", 
        "nutritional", 
        "nutrition assistance",
        "nutrition incentives",
        "nutrition incentive",        
        "food desert",
        "food access",
        "farmers market",
        "snap nutrition",
        "wic nutrition",
        "cafeteria",
        "snack",
        # Query 2 - soda tax, farm to school, labeling, active living
        "beverage tax",
        "sweetened beverages",
        "sugary beverages",
        "soda tax",
        "farm to school",
        "food advertising",
        "food marketing",
        "food labeling",
        "menu labeling",
        "food packaging",
        "food additives",
        "food dyes",
        "active living",
        # Query 3 - obesity prevention, standards, weight loss, healthcare
        "obesity prevention",
        "food standards",
        "weight loss",
        "healthy eating",
        "nutrition counseling",
        "obesity treatment",
        "healthy food",
        "healthy vending",
        "nutrition",
        # Extra
        "food", 
        "milk", 
        "school lunch", 
        "SNAP", 
        "supplemental nutrition assistance",
        "WIC", 
        "farm to school", 
        "restaurant", 
        "grocery", 
        "healthy", 
        "fruit", 
        "vegetable", 
        "meals", 
        "hunger"
        "diet", 
        "dietary", 
    },
    key=len,
    reverse=True,
)


def build_keyword_pattern(keywords: Sequence[str]) -> re.Pattern:
    """Build a compiled regex pattern that matches any keyword (case-insensitive)."""
    escaped = [re.escape(keyword) for keyword in keywords if keyword]
    if not escaped:
        raise ValueError("Keyword list is empty.")
    return re.compile("|".join(escaped), re.IGNORECASE)


def parse_sheet_identifier(raw: str | None) -> int | str | None:
    """Convert a CLI sheet identifier into an int (index) or str (name)."""
    if raw is None:
        return 0  # pandas defaults to first sheet

    try:
        return int(raw)
    except (TypeError, ValueError):
        stripped = str(raw).strip()
        return stripped if stripped else 0


def locate_summary_column(columns: Iterable[str], requested: str) -> str:
    """Find the actual summary column name, matching case-insensitively."""
    lower_map = {col.lower(): col for col in columns}
    key = requested.lower()
    if key in lower_map:
        return lower_map[key]
    raise KeyError(
        f"Column '{requested}' not found. Available columns: {', '.join(columns)}"
    )


def apply_related_flag(
    df: pd.DataFrame, summary_column: str, pattern: re.Pattern
) -> pd.DataFrame:
    """Compute the related flag column using the compiled keyword pattern."""
    summaries = df[summary_column].tolist()
    total = len(summaries)
    if total == 0:
        related_series = pd.Series(dtype=int, index=df.index)
    else:
        bar_width = 30
        last_percent = -1
        related_values = []
        for index, value in enumerate(summaries, start=1):
            if isinstance(value, str):
                match = bool(pattern.search(value))
            elif pd.isna(value):
                match = False
            else:
                match = bool(pattern.search(str(value)))

            related_values.append(int(match))

            percent = int(index * 100 / total)
            if percent != last_percent:
                filled = percent * bar_width // 100
                bar = "#" * filled + "-" * (bar_width - filled)
                print(
                    f"\rKeyword scan progress: {percent:3d}% [{bar}]",
                    end="",
                    flush=True,
                )
                last_percent = percent

        print()  # move to the next line after the progress bar completes
        related_series = pd.Series(related_values, index=df.index, dtype=int)

    if "related" in df.columns:
        df = df.drop(columns=["related"])

    insert_at = df.columns.get_loc(summary_column) + 1
    df.insert(insert_at, "related", related_series)
    return df


def determine_output_path(input_path: Path, output_arg: str | None) -> Path:
    """Resolve the final output path, defaulting to in-place updates."""
    if output_arg:
        return Path(output_arg).expanduser().resolve()

    return input_path


def resolve_sheet_name(workbook, sheet_identifier: int | str | None) -> str:
    """Map a sheet identifier to the canonical sheet name in the workbook."""
    sheet_names = workbook.sheetnames
    if sheet_identifier is None:
        return sheet_names[0]

    if isinstance(sheet_identifier, int):
        return sheet_names[sheet_identifier]

    if sheet_identifier in sheet_names:
        return sheet_identifier

    lowered = sheet_identifier.lower()
    for name in sheet_names:
        if name.lower() == lowered:
            return name

    raise KeyError(
        f"Sheet '{sheet_identifier}' not found. Available sheets: {', '.join(sheet_names)}"
    )


def write_related_column(
    source_path: Path,
    destination_path: Path,
    sheet_identifier: int | str | None,
    summary_column: str,
    related_values: Sequence[int],
) -> None:
    """Update or insert the `related` column in the workbook."""
    workbook = load_workbook(source_path)
    sheet_name = resolve_sheet_name(workbook, sheet_identifier)
    worksheet = workbook[sheet_name]

    header_row = next(
        worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None
    )
    if not header_row:
        raise ValueError(f"Sheet '{sheet_name}' has no header row to inspect.")

    header_lookup = {
        str(value).strip().lower(): index
        for index, value in enumerate(header_row, start=1)
        if value is not None
    }

    summary_key = summary_column.lower()
    if summary_key not in header_lookup:
        raise KeyError(
            f"Summary column '{summary_column}' not found in sheet '{sheet_name}'."
        )

    if "related" in header_lookup:
        related_col_index = header_lookup["related"]
    else:
        related_col_index = header_lookup[summary_key] + 1
        worksheet.insert_cols(related_col_index)
        worksheet.cell(row=1, column=related_col_index, value="related")

    total_rows = len(related_values)
    for row_offset, value in enumerate(related_values, start=2):
        worksheet.cell(row=row_offset, column=related_col_index, value=int(value))

    max_row = worksheet.max_row or 1
    if total_rows + 1 < max_row:
        for row_index in range(total_rows + 2, max_row + 1):
            worksheet.cell(row=row_index, column=related_col_index, value=None)

    workbook.save(destination_path)


def read_excel_single_sheet(
    path: Path, sheet: int | str | None
) -> pd.DataFrame:
    """Read a single sheet from an Excel workbook."""
    data = pd.read_excel(path, sheet_name=sheet)
    if isinstance(data, dict):
        raise ValueError(
            "Multiple sheets were returned; please specify --sheet explicitly."
        )
    return data


def process_file(
    input_path: Path,
    output_path: Path,
    sheet: int | str | None,
    summary_column: str,
    overwrite_input: bool,
) -> None:
    """Orchestrate keyword matching, bill annotation, and file writing."""
    pattern = build_keyword_pattern(KEYWORDS)

    df = read_excel_single_sheet(input_path, sheet)
    actual_summary_col = locate_summary_column(df.columns, summary_column)
    print("Scanning summaries for nutrition keywords...")
    df = apply_related_flag(df, actual_summary_col, pattern)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_related_column(
        source_path=input_path,
        destination_path=output_path,
        sheet_identifier=sheet,
        summary_column=actual_summary_col,
        related_values=df["related"].tolist(),
    )

    total = len(df)
    positives = int(df["related"].sum())

    print(
        f"Annotated {total} records ({positives} marked as related) "
        f"using {len(KEYWORDS)} keywords."
    )
    note = " (overwriting input file)" if overwrite_input else ""
    print(f"Saved updated data to {output_path}{note}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Mark nutrition-related bills based on keyword matches."
    )
    parser.add_argument(
        "excel_path",
        help="Path to the Excel file containing bill records.",
    )
    parser.add_argument(
        "--output",
        help="Optional path for the annotated Excel file; defaults to overwriting the input file.",
    )
    parser.add_argument(
        "--sheet",
        help="Excel sheet index or name (defaults to the first sheet).",
    )
    parser.add_argument(
        "--summary-column",
        default="summary",
        help="Column name that contains bill summaries (default: 'summary').",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    input_path = Path(args.excel_path).expanduser().resolve()
    if not input_path.exists():
        parser.error(f"Input Excel file not found: {input_path}")

    output_path = determine_output_path(input_path, args.output)
    sheet_identifier = parse_sheet_identifier(args.sheet)

    overwriting_input = output_path == input_path
    if overwriting_input:
        print("No output path provided; the original workbook will be overwritten.")
    else:
        print(f"Annotated workbook will be written to {output_path}.")

    try:
        process_file(
            input_path=input_path,
            output_path=output_path,
            sheet=sheet_identifier,
            summary_column=args.summary_column,
            overwrite_input=overwriting_input,
        )
    except Exception as exc:  # noqa: BLE001 - surface clear CLI errors
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()

