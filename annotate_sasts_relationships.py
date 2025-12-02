#!/usr/bin/env python3
"""
Annotate an Excel workbook with LegiScan SAST relationships.

Given the parser output NDJSON (see parser_legiscan_to_schema.py) and a target
worksheet, this script adds or updates two columns:
  * sameAs      → SAST entries with type_id == 1
  * crossfiled  → SAST entries with type_id == 5

Each populated cell contains a JSON array of related bill numbers (and their
LegiScan bill_ids when available). Empty relationships leave the cell blank.
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:  # pragma: no cover - import guard
    print(
        "ERROR: openpyxl is required. Install it with `pip install openpyxl`.",
        file=sys.stderr,
    )
    raise


SAME_AS_TYPE_ID = 1
CROSSFILED_TYPE_ID = 5


@dataclass
class RelationBuckets:
    same_as: List[str]
    crossfiled: List[str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Append SAST relationship columns to a workbook sheet."
    )
    parser.add_argument(
        "--ndjson",
        required=True,
        help="Path to the parser NDJSON file (e.g., data/parsed/bills.ndjson).",
    )
    parser.add_argument(
        "--workbook",
        required=True,
        help="Path to the Excel workbook to update.",
    )
    parser.add_argument(
        "--sheet",
        required=True,
        help="Worksheet name containing the bill rows.",
    )
    parser.add_argument(
        "--id-column",
        default="id",
        help="Header name used to match rows to NDJSON entries (default: id).",
    )
    parser.add_argument(
        "--sameas-column",
        default="sameAs",
        help="Header name for the same-as column (default: sameAs).",
    )
    parser.add_argument(
        "--crossfiled-column",
        default="crossfiled",
        help="Header name for the crossfiled column (default: crossfiled).",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Optional output workbook path. Defaults to in-place overwrite.",
    )
    return parser.parse_args()


def normalize_bill_identifier(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def unique_preserving_order(items: Iterable[str]) -> List[str]:
    seen = set()
    ordered: List[str] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            ordered.append(item)
    return ordered


def build_relation_string(entry: dict) -> Optional[str]:
    bill_number = entry.get("sast_bill_number")
    bill_id = entry.get("sast_bill_id")
    if not bill_number and bill_id is None:
        return None

    number_str = str(bill_number) if bill_number else None
    id_str = None
    if bill_id is not None:
        try:
            id_str = str(int(bill_id))
        except (ValueError, TypeError):
            id_str = str(bill_id)

    if number_str and id_str:
        return f"{number_str} (id={id_str})"
    if number_str:
        return number_str
    return id_str


def collect_relations(ndjson_path: Path) -> Dict[str, RelationBuckets]:
    relations: Dict[str, RelationBuckets] = {}
    with ndjson_path.open("r", encoding="utf-8") as handle:
        for line_number, raw_line in enumerate(handle, start=1):
            line = raw_line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                print(
                    f"[warn] skipping malformed JSON on line {line_number}",
                    file=sys.stderr,
                )
                continue

            bill_id = normalize_bill_identifier(record.get("id") or record.get("bill_id"))
            if not bill_id:
                continue

            raw_sasts = record.get("sasts") or []
            if not isinstance(raw_sasts, list):
                continue

            same_as: List[str] = []
            crossfiled: List[str] = []

            for entry in raw_sasts:
                if not isinstance(entry, dict):
                    continue
                relation_repr = build_relation_string(entry)
                if not relation_repr:
                    continue
                type_id = entry.get("type_id")
                if type_id == SAME_AS_TYPE_ID:
                    same_as.append(relation_repr)
                elif type_id == CROSSFILED_TYPE_ID:
                    crossfiled.append(relation_repr)

            if same_as or crossfiled:
                relations[bill_id] = RelationBuckets(
                    same_as=unique_preserving_order(same_as),
                    crossfiled=unique_preserving_order(crossfiled),
                )
    return relations


def find_column_index(sheet: Worksheet, header_name: str) -> Optional[int]:
    target = header_name.strip().lower()
    for idx, cell in enumerate(sheet[1], start=1):
        value = cell.value
        if value is None:
            continue
        if str(value).strip().lower() == target:
            return idx
    return None


def ensure_column(sheet: Worksheet, header_name: str) -> int:
    existing = find_column_index(sheet, header_name)
    if existing is not None:
        return existing
    new_col = sheet.max_column + 1
    sheet.cell(row=1, column=new_col, value=header_name)
    return new_col


def format_cell_value(values: List[str]) -> Optional[str]:
    if not values:
        return None
    return json.dumps(values, ensure_ascii=False)


def annotate_workbook(
    workbook_path: Path,
    sheet_name: str,
    id_column: str,
    sameas_header: str,
    crossfiled_header: str,
    relations: Dict[str, RelationBuckets],
    output_path: Optional[Path] = None,
) -> None:
    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {workbook_path}")

    ws = wb[sheet_name]

    id_col_idx = find_column_index(ws, id_column)
    if id_col_idx is None:
        raise ValueError(
            f"Column '{id_column}' not found in sheet '{sheet_name}'. "
            "Specify --id-column if the header name differs."
        )

    same_idx = ensure_column(ws, sameas_header)
    cross_idx = ensure_column(ws, crossfiled_header)

    rows_with_relations = 0
    missing_in_ndjson = 0
    for row_idx in range(2, ws.max_row + 1):
        raw_id = ws.cell(row=row_idx, column=id_col_idx).value
        bill_id = normalize_bill_identifier(raw_id)
        if not bill_id:
            continue

        buckets = relations.get(bill_id)
        if buckets is None:
            missing_in_ndjson += 1
            ws.cell(row=row_idx, column=same_idx, value=None)
            ws.cell(row=row_idx, column=cross_idx, value=None)
            continue

        same_val = format_cell_value(buckets.same_as)
        cross_val = format_cell_value(buckets.crossfiled)
        ws.cell(row=row_idx, column=same_idx, value=same_val)
        ws.cell(row=row_idx, column=cross_idx, value=cross_val)
        rows_with_relations += 1

    destination = output_path or workbook_path
    if output_path:
        destination.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destination)

    print(
        f"[ok] annotated sheet '{sheet_name}' in {destination} "
        f"(rows with relations: {rows_with_relations}, "
        f"rows without NDJSON match: {missing_in_ndjson})"
    )


def main() -> None:
    args = parse_args()

    ndjson_path = Path(args.ndjson)
    workbook_path = Path(args.workbook)

    if not ndjson_path.is_file():
        raise FileNotFoundError(f"NDJSON file not found: {ndjson_path}")
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    relations = collect_relations(ndjson_path)
    if not relations:
        print(
            "[warn] No same-as or crossfiled relationships found in NDJSON.",
            file=sys.stderr,
        )

    output_path = Path(args.output) if args.output else None

    annotate_workbook(
        workbook_path=workbook_path,
        sheet_name=args.sheet,
        id_column=args.id_column,
        sameas_header=args.sameas_column,
        crossfiled_header=args.crossfiled_column,
        relations=relations,
        output_path=output_path,
    )


if __name__ == "__main__":
    main()


